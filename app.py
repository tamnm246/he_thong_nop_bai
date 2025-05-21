import os
import sqlite3
from flask import Flask, render_template, request, redirect, url_for, session, flash, make_response, g, send_file
from werkzeug.utils import secure_filename
from datetime import datetime
import openpyxl
from io import BytesIO

# Google Drive API imports
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import io # Cần cho việc download từ Drive

# Thêm logging cơ bản ngay từ đầu để có thể ghi log quá trình khởi tạo
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]')

app = Flask(__name__)
app.secret_key = 'fpt_education_gdktpl_submission_system_!@2025%&*()_secret_drive_v3' # Cập nhật key nếu muốn
app.config['DATABASE'] = 'submissions.db'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit

# Google Drive Configuration
SERVICE_ACCOUNT_FILE = 'service_account.json'
SCOPES = ['https://www.googleapis.com/auth/drive']
DRIVE_FOLDER_ID = '1KtchonCKTyFFAm_olNbcImkwmvsElkMH' # THAY THẾ BẰNG ID THƯ MỤC CỦA BẠN NẾU KHÁC


# Thông tin tài khoản giáo viên
TEACHER_ACCOUNTS = {
    "tamnm": "Fpt28092003@",
    "tupth": "tupth@2025"
}
TEACHER_NAMES = {
    "tamnm": "Nguyễn Minh Tâm",
    "tupth": "Phạm Thị Hồng Tú"
}

# --- Google Drive Service Helper ---
def get_drive_service():
    try:
        if not os.path.exists(SERVICE_ACCOUNT_FILE):
            app.logger.error(f"File '{SERVICE_ACCOUNT_FILE}' không tìm thấy. Google Drive sẽ không hoạt động.")
            return None
        creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('drive', 'v3', credentials=creds)
        return service
    except Exception as e:
        app.logger.error(f"Lỗi khi kết nối Google Drive: {e}")
        return None

# --- Database Helper Functions ---
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(app.config['DATABASE'])
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    with app.app_context():
        db = get_db()
        with app.open_resource('schema.sql', mode='r') as f:
            db.cursor().executescript(f.read())
        db.commit()

def query_db(query, args=(), one=False):
    cur = get_db().execute(query, args)
    rv = cur.fetchall()
    cur.close()
    return (rv[0] if rv else None) if one else rv

def execute_db(query, args=()):
    db = get_db()
    cur = db.cursor()
    cur.execute(query, args)
    db.commit()
    last_id = cur.lastrowid
    cur.close()
    return last_id

@app.cli.command('initdb')
def initdb_command():
    """Initializes the database."""
    init_db()
    print('Initialized the database.')

# --- Routes ---

@app.route('/')
def student_index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit_assignment():
    if 'file-san-pham' not in request.files:
        flash('Không có phần tệp nào được chọn', 'error')
        return redirect(url_for('student_index'))

    student_name = request.form.get('ho-ten-nop', '').strip()
    student_class = request.form.get('lop-nop', '').strip()
    file_to_upload = request.files['file-san-pham']

    if not student_name or not student_class:
        flash('Vui lòng nhập đầy đủ Họ Tên và Lớp.', 'error')
        return redirect(url_for('student_index'))

    if file_to_upload.filename == '':
        flash('Chưa chọn tệp nào để tải lên.', 'error')
        return redirect(url_for('student_index'))

    original_filename = secure_filename(file_to_upload.filename)
    temp_file_path = None 

    if file_to_upload:
        drive_service = get_drive_service()
        if not drive_service:
            flash('Không thể kết nối tới Google Drive. Vui lòng thử lại sau hoặc liên hệ quản trị viên.', 'error')
            return redirect(url_for('student_index'))

        try:
            temp_upload_dir = "temp_uploads"
            if not os.path.exists(temp_upload_dir):
                os.makedirs(temp_upload_dir)
            
            unique_temp_suffix = datetime.now().strftime("%Y%m%d%H%M%S%f")
            temp_filename = f"{unique_temp_suffix}_{original_filename}"
            temp_file_path = os.path.join(temp_upload_dir, temp_filename)
            
            file_to_upload.save(temp_file_path)

            drive_filename = f"{student_class}_{student_name}_{original_filename}" # Tên file trên Drive
            
            file_metadata = {
                'name': drive_filename,
                'parents': [DRIVE_FOLDER_ID]
            }
            media = MediaFileUpload(temp_file_path, resumable=True)
            
            uploaded_file = drive_service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            
            drive_file_id = uploaded_file.get('id')

            if drive_file_id:
                execute_db(
                    "INSERT INTO submissions (student_name, student_class, original_filename, drive_file_id) VALUES (?, ?, ?, ?)",
                    (student_name, student_class, original_filename, drive_file_id)
                )
                flash(f'Bài của bạn ({original_filename}) đã được nộp thành công lên Google Drive!', 'success')
            else:
                flash('Không thể tải file lên Google Drive. ID file không được trả về.', 'error')

        except Exception as e:
            app.logger.error(f"Lỗi khi tải file lên Google Drive hoặc lưu DB: {e}")
            flash(f'Đã xảy ra lỗi trong quá trình nộp bài. Lỗi: {str(e)}', 'error')
        
        finally:
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except Exception as e_remove:
                    app.logger.error(f"Lỗi khi xóa file tạm {temp_file_path}: {e_remove}")
        
        return redirect(url_for('student_index'))

    flash('Đã xảy ra lỗi không xác định khi tải tệp lên.', 'error')
    return redirect(url_for('student_index'))


@app.route('/lookup', methods=['POST'])
def lookup_submission():
    ho_ten = request.form.get('ho-ten-tra-cuu', '').strip()
    lop = request.form.get('lop-tra-cuu', '').strip()
    results = []
    if ho_ten and lop:
        submissions_data = query_db(
            "SELECT original_filename, upload_timestamp FROM submissions WHERE student_name = ? AND student_class = ? ORDER BY upload_timestamp DESC",
            (ho_ten, lop)
        )
        if submissions_data:
            results = submissions_data
            flash(f"Tìm thấy {len(results)} bài nộp cho học sinh {ho_ten} - lớp {lop}.", "info")
        else:
            flash(f"Không tìm thấy bài nộp nào của học sinh {ho_ten} - lớp {lop}.", "warning")
    else:
        flash("Vui lòng nhập đầy đủ Họ tên và Lớp để tra cứu.", "error")
    
    return render_template('index.html', lookup_results=results, prev_ho_ten=ho_ten, prev_lop=lop)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'teacher_user' in session:
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username in TEACHER_ACCOUNTS and TEACHER_ACCOUNTS[username] == password:
            session['teacher_user'] = username
            session['teacher_name'] = TEACHER_NAMES.get(username, username)
            flash('Đăng nhập thành công!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Tên đăng nhập hoặc mật khẩu không đúng.', 'login_error') 
            return render_template('login.html', username=username)

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('teacher_user', None)
    session.pop('teacher_name', None)
    flash('Bạn đã đăng xuất.', 'info')
    return redirect(url_for('login'))

# ĐỊNH NGHĨA HÀM admin_dashboard CHỈ MỘT LẦN TẠI ĐÂY
@app.route('/admin/dashboard')
def admin_dashboard():
    if 'teacher_user' not in session:
        flash('Vui lòng đăng nhập để truy cập trang này.', 'warning')
        return redirect(url_for('login'))

    raw_submissions_data = query_db("SELECT id, student_name, student_class, original_filename, upload_timestamp, drive_file_id FROM submissions ORDER BY upload_timestamp DESC")
    
    formatted_submissions_data = []
    if raw_submissions_data:
        for row in raw_submissions_data:
            submission = dict(row)
            try:
                current_timestamp = submission['upload_timestamp']
                if isinstance(current_timestamp, str):
                    dt_obj = datetime.strptime(current_timestamp.split('.')[0], '%Y-%m-%d %H:%M:%S')
                elif isinstance(current_timestamp, datetime):
                    dt_obj = current_timestamp
                else:
                    dt_obj = None
                    app.logger.warning(f"Dữ liệu upload_timestamp không hợp lệ: {current_timestamp} cho ID: {submission.get('id', 'N/A')}")

                if dt_obj:
                    submission['formatted_upload_timestamp'] = dt_obj.strftime('%d-%m-%Y %H:%M:%S')
                else:
                    submission['formatted_upload_timestamp'] = "N/A"
            
            except ValueError as ve:
                app.logger.error(f"Lỗi định dạng ngày giờ cho submission ID {submission.get('id', 'N/A')}: {ve}. Giá trị gốc: {submission.get('upload_timestamp', 'N/A')}")
                submission['formatted_upload_timestamp'] = str(submission.get('upload_timestamp', 'N/A'))

            formatted_submissions_data.append(submission)

    return render_template('quan_ly.html', submissions=formatted_submissions_data, teacher_name=session.get('teacher_name'))


@app.route('/admin/download_file/<int:submission_id>')
def download_file(submission_id):
    if 'teacher_user' not in session:
        flash('Phiên đăng nhập hết hạn hoặc không hợp lệ.', 'warning')
        return redirect(url_for('login'))

    submission = query_db("SELECT drive_file_id, original_filename FROM submissions WHERE id = ?", (submission_id,), one=True)
    if not submission or not submission['drive_file_id']:
        flash('Không tìm thấy thông tin tệp trên Google Drive cho bài nộp này.', 'error')
        return redirect(url_for('admin_dashboard'))

    drive_service = get_drive_service()
    if not drive_service:
        flash('Không thể kết nối tới Google Drive để tải tệp.', 'error')
        return redirect(url_for('admin_dashboard'))

    try:
        file_id = submission['drive_file_id']
        request_download = drive_service.files().get_media(fileId=file_id)
        file_content_io = io.BytesIO()
        downloader = MediaIoBaseDownload(file_content_io, request_download)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            # app.logger.info(f"Download {int(status.progress() * 100)}%.")
        
        file_content_io.seek(0)
        
        file_metadata = drive_service.files().get(fileId=file_id, fields='mimeType').execute()
        mime_type = file_metadata.get('mimeType', 'application/octet-stream')

        return send_file(
            file_content_io,
            mimetype=mime_type,
            as_attachment=True,
            download_name=submission['original_filename']
        )

    except Exception as e:
        app.logger.error(f"Lỗi khi tải file từ Google Drive (ID: {submission.get('drive_file_id', 'N/A')}): {e}")
        flash(f'Lỗi khi tải tệp "{submission["original_filename"]}" từ Google Drive: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete_submission/<int:submission_id>', methods=['POST'])
def delete_submission(submission_id):
    if 'teacher_user' not in session:
        flash('Phiên đăng nhập hết hạn hoặc không hợp lệ.', 'warning')
        return redirect(url_for('login'))

    submission = query_db("SELECT drive_file_id, original_filename FROM submissions WHERE id = ?", (submission_id,), one=True)
    if not submission:
        flash('Không tìm thấy bài nộp để xóa.', 'error')
        return redirect(url_for('admin_dashboard'))

    drive_service = get_drive_service()
    drive_file_id = submission['drive_file_id']
    original_filename = submission['original_filename']

    try:
        if drive_file_id and drive_service:
            try:
                drive_service.files().delete(fileId=drive_file_id).execute()
                app.logger.info(f"Đã xóa file {drive_file_id} ({original_filename}) trên Google Drive.")
            except Exception as e_drive:
                app.logger.error(f"Lỗi khi xóa file trên Google Drive {drive_file_id}: {e_drive}")
                flash(f'Lưu ý: Gặp lỗi khi xóa tệp "{original_filename}" trên Google Drive. Vui lòng kiểm tra thủ công. Lỗi Drive: {str(e_drive)}', 'warning')
        
        execute_db("DELETE FROM submissions WHERE id = ?", (submission_id,))
        flash(f'Đã xóa thành công bài nộp "{original_filename}" khỏi hệ thống.', 'success')
    
    except Exception as e:
        app.logger.error(f"Lỗi khi xóa bài nộp (DB): {e}")
        flash(f'Lỗi khi xóa bài nộp "{original_filename}" khỏi cơ sở dữ liệu: {str(e)}', 'error')
    
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/export_excel')
def export_excel():
    if 'teacher_user' not in session:
        flash('Phiên đăng nhập hết hạn hoặc không hợp lệ.', 'warning')
        return redirect(url_for('login'))

    submissions_data = query_db("SELECT student_name, student_class, original_filename, upload_timestamp, drive_file_id FROM submissions ORDER BY student_class, student_name, upload_timestamp DESC")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Danh sách nộp bài"

    headers = ["STT", "Họ và Tên", "Lớp", "Tên File", "Thời Gian Nộp", "Google Drive File ID"]
    sheet.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

    for idx, row_data in enumerate(submissions_data):
        try:
            current_timestamp = row_data['upload_timestamp']
            if isinstance(current_timestamp, str):
                timestamp_obj = datetime.strptime(current_timestamp.split('.')[0], '%Y-%m-%d %H:%M:%S')
            elif isinstance(current_timestamp, datetime):
                 timestamp_obj = current_timestamp
            else:
                timestamp_obj = datetime.now() # Fallback
                app.logger.warning(f"Timestamp không đúng định dạng cho export Excel, ID: {row_data.get('id','N/A')}, giá trị: {current_timestamp}")
            
            formatted_timestamp = timestamp_obj.strftime('%d-%m-%Y %H:%M:%S')
        except ValueError as ve:
            app.logger.error(f"Lỗi parse timestamp (Excel): {ve} cho giá trị: {row_data.get('upload_timestamp','N/A')}")
            formatted_timestamp = str(row_data.get('upload_timestamp','N/A'))

        sheet.append([
            idx + 1,
            row_data['student_name'],
            row_data['student_class'],
            row_data['original_filename'],
            formatted_timestamp,
            row_data['drive_file_id'] if row_data['drive_file_id'] else "N/A"
        ])
    
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 3) if max_length < 45 else 45
        sheet.column_dimensions[column].width = adjusted_width

    excel_io = BytesIO()
    workbook.save(excel_io)
    excel_io.seek(0)

    response = make_response(excel_io.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=danh_sach_nop_bai_GD_KTPL_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

if __name__ == '__main__':
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        app.logger.warning(f"CẢNH BÁO: File '{SERVICE_ACCOUNT_FILE}' không tìm thấy. Chức năng Google Drive sẽ KHÔNG hoạt động.")
    
    temp_upload_dir_main = "temp_uploads"
    if not os.path.exists(temp_upload_dir_main):
        try:
            os.makedirs(temp_upload_dir_main)
            app.logger.info(f"Đã tạo thư mục '{temp_upload_dir_main}'.")
        except OSError as e_mkdir:
            app.logger.error(f"Không thể tạo thư mục '{temp_upload_dir_main}': {e_mkdir}")
            
    app.run(debug=True, host='0.0.0.0', port=5000)